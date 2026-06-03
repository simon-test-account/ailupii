"""
Sammenlign hunder.xlsx og kull.xlsx.

Viser hvilke hunder som finnes i den ene filen, men ikke i den andre.
Nyttig for å oppdage feilskrivninger eller hunder som mangler en plass.

Bruk:
    python sammenlign.py
"""

from openpyxl import load_workbook
from pathlib import Path

# === KONFIGURASJON ===
HUNDER_EXCEL = "hunder.xlsx"
KULL_EXCEL = "kull.xlsx"
HUNDER_ARK = "hunder"
KULL_ARK = "Ark1"
HUNDER_OVERSKRIFT_RAD = 2
KULL_OVERSKRIFT_RAD = 3


def les_hunder(filsti):
    """Henter alle filnavn fra hunder.xlsx."""
    if not Path(filsti).exists():
        print(f"FEIL: Finner ikke {filsti}")
        return set()
    
    wb = load_workbook(filsti, data_only=True)
    ws = wb[HUNDER_ARK]
    
    hunder = set()
    for rad in ws.iter_rows(min_row=HUNDER_OVERSKRIFT_RAD + 1, values_only=True):
        if rad[0]:
            navn = str(rad[0]).strip().lower()
            if navn:
                hunder.add(navn)
    return hunder


def les_kull_referanser(filsti):
    """Henter alle filnavn-referanser fra morFilnavn, farFilnavn og valp1-9."""
    if not Path(filsti).exists():
        print(f"FEIL: Finner ikke {filsti}")
        return set(), {}
    
    wb = load_workbook(filsti, data_only=True)
    ws = wb[KULL_ARK]
    
    overskrifter = [c.value for c in ws[KULL_OVERSKRIFT_RAD]]
    
    # Finn kolonner som inneholder hund-referanser
    hund_kolonner = []
    for i, o in enumerate(overskrifter):
        if o is None:
            continue
        o_str = str(o)
        if o_str in ("morFilnavn", "farFilnavn") or o_str.startswith("valp"):
            hund_kolonner.append((i, o_str))
    
    referanser = set()
    referanse_steder = {}  # navn -> liste over (kull_filnavn, kolonne)
    
    for rad in ws.iter_rows(min_row=KULL_OVERSKRIFT_RAD + 1, values_only=True):
        kull_filnavn = str(rad[0]).strip() if rad[0] else "(ukjent)"
        
        for i, kol in hund_kolonner:
            if i < len(rad) and rad[i]:
                navn = str(rad[i]).strip().lower()
                if navn:
                    referanser.add(navn)
                    referanse_steder.setdefault(navn, []).append((kull_filnavn, kol))
    
    return referanser, referanse_steder


def main():
    print("=" * 60)
    print("SAMMENLIGNING AV hunder.xlsx OG kull.xlsx")
    print("=" * 60)
    
    hunder = les_hunder(HUNDER_EXCEL)
    referanser, steder = les_kull_referanser(KULL_EXCEL)
    
    if not hunder and not referanser:
        return
    
    print(f"\nHunder i hunder.xlsx: {len(hunder)}")
    print(f"Hunder referert i kull.xlsx: {len(referanser)}")
    
    kun_i_hunder = hunder - referanser
    kun_i_kull = referanser - hunder
    i_begge = hunder & referanser
    
    print(f"I begge filer: {len(i_begge)}")
    
    # ----- KUN I HUNDER -----
    print("\n" + "=" * 60)
    print(f"HUNDER UTEN KULLREFERANSE ({len(kun_i_hunder)})")
    print("-" * 60)
    print("Disse er registrert som hunder, men er ikke nevnt som")
    print("mor/far/valp i noe kull. Helt OK - kanskje fôrhunder eller")
    print("hunder uten kull enda.")
    print()
    if kun_i_hunder:
        for h in sorted(kun_i_hunder):
            print(f"  {h}")
    else:
        print("  (ingen)")
    
    # ----- KUN I KULL -----
    print("\n" + "=" * 60)
    print(f"HUNDER REFERERT, MEN IKKE REGISTRERT ({len(kun_i_kull)})")
    print("-" * 60)
    print("Disse er nevnt i et kull, men finnes ikke i hunder.xlsx.")
    print("Sjekk om de bør legges til, eller om det er en skrivefeil.")
    print()
    if kun_i_kull:
        for h in sorted(kun_i_kull):
            forekomster = steder.get(h, [])
            print(f"  {h}")
            for kull_navn, kol in forekomster:
                print(f"      -> {kull_navn} ({kol})")
    else:
        print("  (ingen)")
    
    print("\n" + "=" * 60)
    print("FERDIG")
    print("=" * 60)


if __name__ == "__main__":
    main()