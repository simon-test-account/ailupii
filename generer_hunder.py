"""
Generer hundeprofil-sider fra Excel-ark.

Bruk:
1. Fyll ut hunder.xlsx (en rad per hund, overskrifter på rad 2)
2. Kjør: python generer_hunder.py
3. HTML-filer lages/oppdateres i samme mappe

Krever: pip install openpyxl

Manuelle endringer:
  Alt du skriver mellom BEVAR_START og BEVAR_SLUTT beholdes ved
  oppdateringer. Skriv galleri, ekstra tekst eller annet der.
"""

from openpyxl import load_workbook
from pathlib import Path
import re

# === KONFIGURASJON ===
EXCEL_FIL = "hunder.xlsx"
ARK_NAVN = "hunder"
OVERSKRIFT_RAD = 2  # rad nummer hvor kolonneoverskriftene står
KENNEL_NAVN = "Kennel Ailupii"
FACEBOOK_URL = "https://facebook.com/dinkennel"

# === BEVAR-MARKØRER ===
BEVAR_START = "<!-- BEVAR_START: ekstra -->"
BEVAR_SLUTT = "<!-- BEVAR_SLUTT: ekstra -->"


# === HJELPEFUNKSJONER ===

def rens(verdi):
    """Fjerner None, tabulator-tegn og trim whitespace."""
    if verdi is None:
        return ""
    return str(verdi).replace("\t", "").strip()


def er_tom(verdi):
    """Sjekk om en verdi er tom (None, tom streng, eller 0)."""
    if verdi is None:
        return True
    if isinstance(verdi, str) and verdi.strip() == "":
        return True
    if verdi == 0:
        return True
    return False


def oversett_kjonn(kode):
    """T -> Tispe, H -> Hannhund, ellers tom."""
    kode = rens(kode).upper()
    if kode == "T":
        return "Tispe"
    if kode == "H":
        return "Hannhund"
    return ""


def lag_meta_info(kjonn, fodt):
    """Lager 'Tispe · Født 2018' eller 'Født 2018' hvis kjønn mangler."""
    kjonn_tekst = oversett_kjonn(kjonn)
    fodt_tekst = rens(fodt)
    
    deler = []
    if kjonn_tekst:
        deler.append(kjonn_tekst)
    if fodt_tekst:
        deler.append(f"Født {fodt_tekst}")
    return " · ".join(deler)


def lag_navn_med_tittel(navn, tittel):
    """'N UCH · Pieni Pilvi av Vintervidda' eller bare navnet."""
    navn = rens(navn)
    tittel = rens(tittel)
    if tittel and navn:
        return f"{tittel} · {navn}"
    return navn


def lag_faktaboks(rad):
    """Bygger tabellrader. HD og øyenlysning droppes hvis tom/0."""
    felter = []
    
    regnavn = rens(rad.get("regnavn"))
    if regnavn:
        felter.append(("Registrert navn", regnavn))


    farge = rens(rad.get("farge"))
    if farge:
        felter.append(("Farge", farge))
    
    if not er_tom(rad.get("hd")):
        felter.append(("HD", rens(rad.get("hd"))))
    
    if not er_tom(rad.get("øyenlysning")):
        felter.append(("Øyenlysning", rens(rad.get("øyenlysning"))))
    
    oppdretter = rens(rad.get("oppdretter"))
    if oppdretter:
        felter.append(("Oppdretter", oppdretter))
    
    eier = rens(rad.get("eier"))
    if eier:
        felter.append(("Eier", eier))
    
    linjer = []
    for label, verdi in felter:
        linjer.append("            <tr>")
        linjer.append(f"                <td>{label}</td>")
        linjer.append(f"                <td>{verdi}</td>")
        linjer.append("            </tr>")
    
    return "\n".join(linjer)


def lag_forelder_detalj(hd, tittel):
    """'HD: A · N UCH' eller bare den ene som finnes."""
    deler = []
    if not er_tom(hd):
        deler.append(f"HD: {rens(hd)}")
    tittel_tekst = rens(tittel)
    if tittel_tekst:
        deler.append(tittel_tekst)
    return " · ".join(deler)


def lag_kull_seksjon(kallenavn, ant_kull):
    """Lager kull-seksjonen med plassholder-lenker. Tom hvis 0/tom."""
    if er_tom(ant_kull):
        return ""
    
    try:
        antall = int(ant_kull)
    except (ValueError, TypeError):
        return ""
    
    if antall <= 0:
        return ""
    
    elementer = []
    for i in range(1, antall + 1):
        elementer.append(f'''        <a href="kull-{i:02d}.html" class="kull-element">
            <p class="kull-navn">Kull {i}</p>
            <p class="kull-detalj">Fyll inn detaljer senere</p>
        </a>''')
    
    elementer_tekst = "\n".join(elementer)
    return f"""
    <h2>Kull etter {kallenavn}</h2>
    <div class="kull-liste">
{elementer_tekst}
    </div>
"""


def hent_bevart_innhold(html_sti):
    """Leser eksisterende fil og henter ut bevart blokk hvis den finnes."""
    if not html_sti.exists():
        return ""
    
    try:
        innhold = html_sti.read_text(encoding="utf-8")
    except Exception:
        return ""
    
    # Finn alt mellom BEVAR_START og BEVAR_SLUTT
    pattern = re.escape(BEVAR_START) + r"(.*?)" + re.escape(BEVAR_SLUTT)
    match = re.search(pattern, innhold, re.DOTALL)
    if match:
        return match.group(1)
    return ""


# === HTML-MAL ===

def bygg_html(rad, bevart_innhold=""):
    """Bygger den komplette HTML-en for en hund."""
    filnavn = rens(rad.get("filnavn"))
    kallenavn = rens(rad.get("kallenavn"))
    regnavn = rens(rad.get("regnavn"))
    
    # Tittel i nettleserfanen
    side_tittel = kallenavn or regnavn or filnavn
    
    # Hovedoverskrift med eventuell tittel foran
    h1_tekst = rens(regnavn)
    
    # Kallenavn-linje (droppes hvis tom)
    tittel = rens(rad.get("tittel"))
    if kallenavn and tittel:
        kallenavn_linje = f'    <p class="kallenavn">«{kallenavn}» · {tittel}</p>'
    elif kallenavn:
        kallenavn_linje = f'    <p class="kallenavn">«{kallenavn}»</p>'
    elif tittel:
        kallenavn_linje = f'    <p class="kallenavn">{tittel}</p>'
    else:
        kallenavn_linje = ""
    
    meta_info = lag_meta_info(rad.get("kjønn"), rad.get("født"))
    faktaboks = lag_faktaboks(rad)
    
    far_detalj = lag_forelder_detalj(rad.get("farHD"), rad.get("farTittel"))
    mor_detalj = lag_forelder_detalj(rad.get("morHD"), rad.get("morTittel"))
    
    kull_seksjon = lag_kull_seksjon(kallenavn or regnavn, rad.get("antKull"))
    
    # Hvis ingen bevart-blokk fra før, sett inn tom plassholder
    if not bevart_innhold:
        bevart_innhold = "\n    "
    
    html = f"""<!DOCTYPE html>
<html lang="nb">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{side_tittel} - {KENNEL_NAVN}</title>
    <link rel="stylesheet" href="style.css">
</head>
<body>

<div class="container">

    <nav>
        <a href="index.html">Forsiden</a>
        <a href="vare-hunder.html" class="aktiv">Våre hunder</a>
        <a href="kull.html">Kull</a>
        <a href="etterkommere.html">Etterkommere</a>
        <a href="om-oss.html">Om oss</a>
        <a href="{FACEBOOK_URL}" class="facebook-lenke" target="_blank" rel="noopener">Facebook</a>
    </nav>

    <p class="meta-info">{meta_info}</p>
    <h1>{h1_tekst}</h1>
{kallenavn_linje}

    <img src="bilder/{filnavn}.jpg" alt="{kallenavn or regnavn}" class="hovedbilde-img">

    <div class="faktaboks">
        <table>
{faktaboks}
        </table>
    </div>

    <h2>Om {kallenavn or regnavn}</h2>
    <p>{rens(rad.get("om hunden"))}</p>

    <h2>Stamtavle</h2>
    <div class="stamtavle">
        <div class="forelder">
            <p class="label">Far</p>
            <p class="navn">{rens(rad.get("farNavn"))}</p>
            <p class="detalj">{far_detalj}</p>
        </div>
        <div class="forelder">
            <p class="label">Mor</p>
            <p class="navn">{rens(rad.get("morNavn"))}</p>
            <p class="detalj">{mor_detalj}</p>
        </div>
    </div>
{kull_seksjon}
    {BEVAR_START}{bevart_innhold}{BEVAR_SLUTT}

    <footer>
        {KENNEL_NAVN} · Følg oss på <a href="{FACEBOOK_URL}" target="_blank" rel="noopener">Facebook</a> for siste nytt
    </footer>

</div>

</body>
</html>
"""
    return html


# === HOVEDLOGIKK ===

def les_excel(filsti):
    """Leser excel og returnerer liste med dict per hund."""
    wb = load_workbook(filsti, data_only=True)
    ws = wb[ARK_NAVN]
    
    overskrifter = [c.value for c in ws[OVERSKRIFT_RAD]]
    
    rader = []
    for rad in ws.iter_rows(min_row=OVERSKRIFT_RAD + 1, values_only=True):
        # Ignorer rad hvis første kolonne (filnavn) er tom
        if er_tom(rad[0]):
            continue
        
        rad_dict = {}
        for overskrift, verdi in zip(overskrifter, rad):
            if overskrift:
                rad_dict[overskrift] = verdi
        rader.append(rad_dict)
    
    return rader


def main():
    if not Path(EXCEL_FIL).exists():
        print(f"FEIL: Finner ikke {EXCEL_FIL}")
        print(f"Plasser filen i samme mappe som dette scriptet.")
        return
    
    print(f"Leser {EXCEL_FIL}...")
    rader = les_excel(EXCEL_FIL)
    print(f"Fant {len(rader)} hund(er) a behandle.\n")
    
    laget = 0
    oppdatert = 0
    
    for rad in rader:
        filnavn = rens(rad.get("filnavn"))
        if not filnavn:
            continue
        
        html_sti = Path(f"{filnavn}.html")
        eksisterer = html_sti.exists()
        
        bevart = hent_bevart_innhold(html_sti)
        html = bygg_html(rad, bevart)
        html_sti.write_text(html, encoding="utf-8")
        
        if eksisterer:
            print(f"  Oppdaterte {html_sti.name}")
            oppdatert += 1
        else:
            print(f"  Laget {html_sti.name}")
            laget += 1
    
    print(f"\nFerdig!")
    print(f"  Nye filer: {laget}")
    print(f"  Oppdaterte filer: {oppdatert}")


if __name__ == "__main__":
    main()
