"""
Endre bildenavn fra gammelt filnavn til nytt filnavn.

Leser mapping fra data_hunder.xlsx (filnavnG -> filnavn) og endrer
filnavn på bilder i bilder/-mappen.

Eksempel: bilder/luna.jpg -> bilder/pieni_pilvi_av_vintervidda.jpg

Bruk:
    python rename_bilder.py           # Vis hva som vil skje (dry-run)
    python rename_bilder.py --gjor    # Utfør faktisk endringen

Krever: pip install openpyxl
"""

from pathlib import Path
import sys
from kennel_felles import HUNDER_EXCEL, HUNDER_ARK, HUNDER_OVERSKRIFT_RAD, rens
from openpyxl import load_workbook

BILDE_MAPPE = "bilder"
GYLDIGE_ENDELSER = {".jpg", ".jpeg", ".png", ".webp", ".gif"}


def hent_mapping():
    """Returnerer dict: filnavnG (gammelt) -> filnavn (nytt)."""
    if not Path(HUNDER_EXCEL).exists():
        print(f"FEIL: Finner ikke {HUNDER_EXCEL}")
        return {}
    
    wb = load_workbook(HUNDER_EXCEL, data_only=True)
    if HUNDER_ARK not in wb.sheetnames:
        print(f"FEIL: Finner ikke arket '{HUNDER_ARK}' i {HUNDER_EXCEL}")
        return {}
    
    ws = wb[HUNDER_ARK]
    overskrifter = [c.value for c in ws[HUNDER_OVERSKRIFT_RAD]]
    
    # Finn kolonneindekser
    try:
        idx_gammel = overskrifter.index("filnavnG")
        idx_ny = overskrifter.index("filnavn")
    except ValueError:
        print("FEIL: Mangler kolonne 'filnavnG' eller 'filnavn' i arket")
        return {}
    
    mapping = {}
    for rad in ws.iter_rows(min_row=HUNDER_OVERSKRIFT_RAD + 1, values_only=True):
        gammel = rens(rad[idx_gammel]).lower() if idx_gammel < len(rad) else ""
        ny = rens(rad[idx_ny]) if idx_ny < len(rad) else ""
        if gammel and ny and gammel != ny.lower():
            mapping[gammel] = ny
    
    return mapping


def finn_bilder():
    """Returnerer liste over bildefiler i bilder-mappen."""
    bilde_dir = Path(BILDE_MAPPE)
    if not bilde_dir.exists():
        print(f"FEIL: Finner ikke mappen '{BILDE_MAPPE}/'")
        return []
    
    bilder = []
    for fil in bilde_dir.iterdir():
        if fil.is_file() and fil.suffix.lower() in GYLDIGE_ENDELSER:
            bilder.append(fil)
    return bilder


def main():
    gjor_endring = "--gjor" in sys.argv
    
    print("=" * 60)
    print("BILDENAVNBYTTE")
    print("=" * 60)
    
    if not gjor_endring:
        print("\nDRY-RUN: Viser hva som VILLE skjedd. Bruk --gjor for å utføre.\n")
    
    mapping = hent_mapping()
    if not mapping:
        return
    
    print(f"Lest mapping for {len(mapping)} hunder fra {HUNDER_EXCEL}")
    
    bilder = finn_bilder()
    if not bilder:
        return
    
    print(f"Fant {len(bilder)} bilder i {BILDE_MAPPE}/\n")
    
    skal_endres = []
    allerede_riktig = []
    ikke_funnet = []
    
    for bilde in bilder:
        # Hent navnet uten endelse, gjør lowercase
        navn_lower = bilde.stem.lower()
        endelse = bilde.suffix.lower()
        
        if navn_lower in mapping:
            nytt_navn = f"{mapping[navn_lower]}{endelse}"
            ny_sti = bilde.parent / nytt_navn
            
            if ny_sti.exists() and ny_sti != bilde:
                print(f"  ADVARSEL: {bilde.name} -> {nytt_navn} (målfilen finnes allerede!)")
                continue
            
            skal_endres.append((bilde, ny_sti))
        elif navn_lower in [v.lower() for v in mapping.values()]:
            # Bildet har allerede et nytt filnavn
            allerede_riktig.append(bilde)
        else:
            ikke_funnet.append(bilde)
    
    # ----- ENDRINGER -----
    print("=" * 60)
    print(f"BILDER SOM SKAL DØPES OM ({len(skal_endres)})")
    print("-" * 60)
    if skal_endres:
        for gammel, ny in skal_endres:
            print(f"  {gammel.name}")
            print(f"    -> {ny.name}")
    else:
        print("  (ingen)")
    
    # ----- ALLEREDE RIKTIG -----
    if allerede_riktig:
        print("\n" + "=" * 60)
        print(f"BILDER MED RIKTIG NAVN ALLEREDE ({len(allerede_riktig)})")
        print("-" * 60)
        for bilde in allerede_riktig:
            print(f"  {bilde.name}")
    
    # ----- IKKE FUNNET -----
    if ikke_funnet:
        print("\n" + "=" * 60)
        print(f"BILDER UTEN MATCH I EXCEL ({len(ikke_funnet)})")
        print("-" * 60)
        print("Disse er ikke nevnt i data_hunder.xlsx og blir ikke endret:")
        for bilde in ikke_funnet:
            print(f"  {bilde.name}")
    
    # ----- UTFØR -----
    print("\n" + "=" * 60)
    if gjor_endring and skal_endres:
        print("UTFØRER ENDRINGER...")
        print("-" * 60)
        feil = 0
        ok = 0
        for gammel, ny in skal_endres:
            try:
                gammel.rename(ny)
                print(f"  OK: {gammel.name} -> {ny.name}")
                ok += 1
            except Exception as e:
                print(f"  FEIL: {gammel.name} - {e}")
                feil += 1
        print(f"\nFerdig! {ok} OK, {feil} feil.")
    elif gjor_endring:
        print("Ingen endringer å utføre.")
    else:
        print(f"Kjør 'python rename_bilder.py --gjor' for å utføre.")


if __name__ == "__main__":
    main()