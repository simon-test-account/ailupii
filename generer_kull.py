"""
Generer kullsider og kulloversikt fra Excel-ark.

Bruk:
1. Fyll ut kull.xlsx (en rad per kull, overskrifter på rad 3)
2. Sørg for at hunder.xlsx er i samme mappe (for å hente foreldre/valper-info)
3. Kjør: python generer_kull.py
4. Oversikt (kull.html) og en kull-XX.html per kull genereres/oppdateres

Krever: pip install openpyxl

Manuelle endringer:
  Alt mellom BEVAR_START og BEVAR_SLUTT beholdes ved oppdateringer.
"""

from openpyxl import load_workbook
from pathlib import Path
import re

# === KONFIGURASJON ===
KULL_EXCEL = "kull.xlsx"
HUNDER_EXCEL = "hunder.xlsx"
KULL_ARK = "Ark1"
HUNDER_ARK = "hunder"
KULL_OVERSKRIFT_RAD = 3
HUNDER_OVERSKRIFT_RAD = 2
KENNEL_NAVN = "Kennel Ailupii"
FACEBOOK_URL = "https://facebook.com/dinkennel"
OVERSIKT_FIL = "kull.html"

# === BEVAR-MARKØRER ===
BEVAR_START = "<!-- BEVAR_START: ekstra -->"
BEVAR_SLUTT = "<!-- BEVAR_SLUTT: ekstra -->"


# === HJELPEFUNKSJONER ===

def rens(verdi):
    if verdi is None:
        return ""
    return str(verdi).replace("\t", "").strip()


def er_tom(verdi):
    if verdi is None:
        return True
    if isinstance(verdi, str) and verdi.strip() == "":
        return True
    if verdi == 0:
        return True
    return False


def til_romertall(n):
    """Konverterer 1-50 til romertall. 0 returnerer 'Nulla'."""
    if n == 0:
        return "Nulla"
    
    verdier = [
        (40, "XL"), (10, "X"), (9, "IX"), (5, "V"), (4, "IV"), (1, "I")
    ]
    resultat = ""
    for verdi, symbol in verdier:
        while n >= verdi:
            resultat += symbol
            n -= verdi
    return resultat


def format_dato_norsk(dato):
    """Hvis dato er en faktisk dato-objekt, formater pent. Ellers returner som tekst."""
    if dato is None:
        return ""
    
    # Hvis det er en faktisk datetime, formater pent
    try:
        from datetime import datetime, date
        if isinstance(dato, (datetime, date)):
            mnd_navn = ["januar", "februar", "mars", "april", "mai", "juni",
                       "juli", "august", "september", "oktober", "november", "desember"]
            return f"{dato.day}. {mnd_navn[dato.month - 1]} {dato.year}"
    except Exception:
        pass
    
    return rens(dato)


# === LESE HUNDER ===

def les_hunder(filsti):
    """Returnerer dict med filnavn -> hund-info."""
    if not Path(filsti).exists():
        print(f"  ADVARSEL: {filsti} finnes ikke. Bruker filnavn som fallback for alle.")
        return {}
    
    wb = load_workbook(filsti, data_only=True)
    ws = wb[HUNDER_ARK]
    
    overskrifter = [c.value for c in ws[HUNDER_OVERSKRIFT_RAD]]
    
    hunder = {}
    for rad in ws.iter_rows(min_row=HUNDER_OVERSKRIFT_RAD + 1, values_only=True):
        if er_tom(rad[0]):
            continue
        rad_dict = {}
        for overskrift, verdi in zip(overskrifter, rad):
            if overskrift:
                rad_dict[overskrift] = verdi
        filnavn = rens(rad_dict.get("filnavn")).lower()
        hunder[filnavn] = rad_dict
    
    return hunder


def oversett_kjonn(kode):
    kode = rens(kode).upper()
    if kode == "T":
        return "Tispe"
    if kode == "H":
        return "Hannhund"
    return ""


# === LESE KULL ===

def les_kull(filsti):
    """Returnerer liste med kull-dict (i excel-rekkefølge)."""
    wb = load_workbook(filsti, data_only=True)
    ws = wb[KULL_ARK]
    
    overskrifter = [c.value for c in ws[KULL_OVERSKRIFT_RAD]]
    
    kull = []
    for rad in ws.iter_rows(min_row=KULL_OVERSKRIFT_RAD + 1, values_only=True):
        if er_tom(rad[0]):
            continue
        
        # Ignorer rader uten navn (kun filnavn fylt ut)
        if er_tom(rad[1]):  # 'navn' er andre kolonne
            continue
        
        rad_dict = {}
        for overskrift, verdi in zip(overskrifter, rad):
            if overskrift:
                rad_dict[overskrift] = verdi
        kull.append(rad_dict)
    
    return kull


# === BEVAR-LOGIKK ===

def hent_bevart_innhold(html_sti):
    if not html_sti.exists():
        return ""
    try:
        innhold = html_sti.read_text(encoding="utf-8")
    except Exception:
        return ""
    pattern = re.escape(BEVAR_START) + r"(.*?)" + re.escape(BEVAR_SLUTT)
    match = re.search(pattern, innhold, re.DOTALL)
    if match:
        return match.group(1)
    return ""


# === HJELPERE FOR HUND-INFO ===

def hent_hund_info(filnavn, hunder):
    """Returnerer (regnavn, kallenavn, kjonn, eier, finnes_i_excel)."""
    filnavn_lower = rens(filnavn).lower()
    if filnavn_lower in hunder:
        hund = hunder[filnavn_lower]
        return (
            rens(hund.get("regnavn")) or filnavn,
            rens(hund.get("kallenavn")) or filnavn,
            oversett_kjonn(hund.get("kjønn")),
            rens(hund.get("eier")),
            True,
        )
    # Fallback - bruk filnavn for alt
    return (filnavn, filnavn, "", "", False)


# === GENERERE OVERSIKT (kull.html) ===

def lag_kullkort_for_oversikt(kull, hunder, er_utlant):
    """Lager ett kullkort til oversikten."""
    filnavn = rens(kull.get("filnavn"))
    navn = rens(kull.get("navn"))
    dato = format_dato_norsk(kull.get("dato"))
    antall = kull.get("antall")
    
    mor_fil = rens(kull.get("morFilnavn"))
    far_fil = rens(kull.get("farFilnavn"))
    
    mor_regnavn, mor_kallenavn, _, mor_eier, _ = hent_hund_info(mor_fil, hunder)
    far_regnavn, far_kallenavn, _, far_eier, _ = hent_hund_info(far_fil, hunder)
    
    # Detaljlinje
    if antall == -1:
        detalj = f"Planlagt {dato}"
    elif er_tom(antall):
        detalj = f"Termin {dato}"
    else:
        detalj = f"Født {dato} · {antall} valper"
    
    # Under foreldrene: kallenavn for egne kull, eier for utlånte
    if er_utlant:
        mor_under = mor_eier
        far_under = far_eier
    else:
        mor_under = f"«{mor_kallenavn}»" if mor_kallenavn else ""
        far_under = f"«{far_kallenavn}»" if far_kallenavn else ""
    
    return f"""        <a href="{filnavn}.html" class="kull-kort">
            <div class="kull-info">
                <p class="kull-navn">{navn}</p>
                <p class="kull-detalj">{detalj}</p>
            </div>
            <div class="foreldre">
                <div class="forelder-bilde">
                    <img src="bilder/{mor_fil}.jpg" alt="{mor_kallenavn}">
                    <div class="forelder-tekst">
                        <p class="forelder-label">Mor</p>
                        <p class="forelder-navn">{mor_regnavn}</p>
                        <p class="forelder-detalj">{mor_under}</p>
                    </div>
                </div>
                <div class="forelder-bilde">
                    <img src="bilder/{far_fil}.jpg" alt="{far_kallenavn}">
                    <div class="forelder-tekst">
                        <p class="forelder-label">Far</p>
                        <p class="forelder-navn">{far_regnavn}</p>
                        <p class="forelder-detalj">{far_under}</p>
                    </div>
                </div>
            </div>
        </a>"""


def generer_oversikt(alle_kull, hunder):
    """Bygger kull.html med egne kull og utlånte kull."""
    egne = [k for k in alle_kull if not rens(k.get("filnavn")).startswith("kull-utlant")]
    utlante = [k for k in alle_kull if rens(k.get("filnavn")).startswith("kull-utlant")]
    
    egne_kort = "\n\n".join(lag_kullkort_for_oversikt(k, hunder, False) for k in egne)
    utlante_kort = "\n\n".join(lag_kullkort_for_oversikt(k, hunder, True) for k in utlante)
    
    # Behold BEVAR fra eksisterende kull.html hvis den finnes
    bevart = hent_bevart_innhold(Path(OVERSIKT_FIL))
    if not bevart:
        bevart = "\n    "
    
    utlante_seksjon = ""
    if utlante_kort:
        utlante_seksjon = f"""
    <h2>Utlånte kull</h2>

    <div class="kull-grid">

{utlante_kort}

    </div>
"""
    
    html = f"""<!DOCTYPE html>
<html lang="nb">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Kull - {KENNEL_NAVN}</title>
    <link rel="stylesheet" href="style.css">
</head>
<body>

<div class="container">

    <nav>
        <a href="index.html">Forsiden</a>
        <a href="vare-hunder.html">Våre hunder</a>
        <a href="kull.html" class="aktiv">Kull</a>
        <a href="etterkommere.html">Etterkommere</a>
        <a href="om-oss.html">Om oss</a>
        <a href="{FACEBOOK_URL}" class="facebook-lenke" target="_blank" rel="noopener">Facebook</a>
    </nav>

    <h1>Kull</h1>
    <h2>Våre kull</h2>

    <div class="kull-grid">

{egne_kort}

    </div>
{utlante_seksjon}
    {BEVAR_START}{bevart}{BEVAR_SLUTT}

    <footer>
        {KENNEL_NAVN} · Følg oss på <a href="{FACEBOOK_URL}" target="_blank" rel="noopener">Facebook</a> for siste nytt
    </footer>

</div>

</body>
</html>
"""
    return html


# === GENERERE INDIVIDUELL KULLSIDE ===

def lag_meta_info_for_kull(filnavn, er_utlant):
    """Returnerer 'Ailupii III' for egne kull, tom for utlånte."""
    if er_utlant:
        return ""
    # filnavn er kull-XX, hent ut tallet
    match = re.match(r"kull-(\d+)", filnavn)
    if not match:
        return ""
    nummer = int(match.group(1))
    return f"Ailupii {til_romertall(nummer)}"


def lag_forelder_kort(filnavn, hunder, er_utlant):
    """Lager et klikkbart forelder-kort til kullsiden."""
    regnavn, kallenavn, _, eier, _ = hent_hund_info(filnavn, hunder)
    
    # Under-tekst: kallenavn for egne kull, eier for utlånte
    if er_utlant:
        under = eier
    else:
        under = f"«{kallenavn}»" if kallenavn and kallenavn != filnavn else ""
    
    return regnavn, kallenavn, under


def lag_valpe_kort(valp_filnavn, hunder):
    """Lager ett valpekort. Returnerer HTML-streng."""
    filnavn = rens(valp_filnavn)
    if not filnavn:
        return ""
    
    filnavn_lower = filnavn.lower()
    if filnavn_lower in hunder:
        hund = hunder[filnavn_lower]
        regnavn = rens(hund.get("regnavn")) or filnavn
        kallenavn = rens(hund.get("kallenavn"))
        kjonn = oversett_kjonn(hund.get("kjønn"))
        
            # Bygg detaljlinje bare med det som faktisk finnes
        deler = []
        if kjonn:
            deler.append(kjonn)
        if kallenavn:
            deler.append(f"«{kallenavn}»")
        kjonn_kallenavn = " · ".join(deler)
        
        # Bilde-alt-tekst: kallenavn hvis finnes, ellers regnavn
        alt_tekst = kallenavn or regnavn
        return f"""        <a href="{filnavn}.html" class="valpe-kort">
            <img src="bilder/{filnavn}.jpg" alt="{alt_tekst}">
            <div class="valpe-info">
                <p class="valpe-navn">{regnavn}</p>
                <p class="valpe-detalj">{kjonn_kallenavn}</p>
            </div>
        </a>"""
    else:
        # Valpen finnes ikke i hunder.xlsx - ingen lenke
        return f"""        <div class="valpe-kort">
            <img src="bilder/{filnavn}.jpg" alt="{filnavn}">
            <div class="valpe-info">
                <p class="valpe-navn">{filnavn}</p>
                <p class="valpe-detalj"></p>
            </div>
        </div>"""


def generer_kullside(kull, hunder):
    """Lager HTML for én individuell kull-XX.html."""
    filnavn = rens(kull.get("filnavn"))
    navn = rens(kull.get("navn"))
    dato = format_dato_norsk(kull.get("dato"))
    antall = kull.get("antall")
    om_kullet = rens(kull.get("Om kullet"))
    
    er_utlant = filnavn.startswith("kull-utlant")
    er_planlagt = er_tom(antall) or antall == -1
    
    # Meta-info (Ailupii III osv.)
    meta_info = lag_meta_info_for_kull(filnavn, er_utlant)
    meta_linje = f'    <p class="meta-info">{meta_info}</p>' if meta_info else ""
    
    # Kallenavn-linje under H1 (her brukes for dato)
    if antall == -1:
        dato_linje = f"Planlagt {dato}"
    elif er_planlagt:
        dato_linje = f"Termin {dato}"
    else:
        dato_linje = f"Født {dato} · {antall} valper"
    
    # Foreldre
    mor_fil = rens(kull.get("morFilnavn"))
    far_fil = rens(kull.get("farFilnavn"))
    
    mor_regnavn, mor_kallenavn, mor_under = lag_forelder_kort(mor_fil, hunder, er_utlant)
    far_regnavn, far_kallenavn, far_under = lag_forelder_kort(far_fil, hunder, er_utlant)
    
    # Valper - sjekk alle kolonner som starter med "valp"
    valp_kort_liste = []
    if not er_planlagt:
        for nokkel, verdi in kull.items():
            if str(nokkel).startswith("valp") and verdi:
                kort = lag_valpe_kort(verdi, hunder)
                if kort:
                    valp_kort_liste.append(kort)
    
    # Valpe-seksjon (kun hvis det finnes valper)
    if valp_kort_liste:
        valper_html = "\n".join(valp_kort_liste)
        valpe_seksjon = f"""
    <h2>Valpene</h2>
    <div class="valpe-grid">

{valper_html}

    </div>
"""
    else:
        valpe_seksjon = ""
    
    # Bevart innhold
    html_sti = Path(f"{filnavn}.html")
    bevart = hent_bevart_innhold(html_sti)
    if not bevart:
        bevart = "\n    "
    
    html = f"""<!DOCTYPE html>
<html lang="nb">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{navn} - {KENNEL_NAVN}</title>
    <link rel="stylesheet" href="style.css">
</head>
<body>

<div class="container">

    <nav>
        <a href="index.html">Forsiden</a>
        <a href="vare-hunder.html">Våre hunder</a>
        <a href="kull.html" class="aktiv">Kull</a>
        <a href="etterkommere.html">Etterkommere</a>
        <a href="om-oss.html">Om oss</a>
        <a href="{FACEBOOK_URL}" class="facebook-lenke" target="_blank" rel="noopener">Facebook</a>
    </nav>

{meta_linje}
    <h1>{navn}</h1>
    <p class="kallenavn">{dato_linje}</p>

    <div class="foreldre-kompakt">
        
        <a href="{mor_fil}.html" class="forelder-kort">
            <img src="bilder/{mor_fil}.jpg" alt="{mor_kallenavn}">
            <div class="forelder-info">
                <p class="forelder-label">Mor</p>
                <p class="forelder-navn">{mor_regnavn}</p>
                <p class="forelder-kallenavn">{mor_under}</p>
            </div>
        </a>
        
        <a href="{far_fil}.html" class="forelder-kort">
            <img src="bilder/{far_fil}.jpg" alt="{far_kallenavn}">
            <div class="forelder-info">
                <p class="forelder-label">Far</p>
                <p class="forelder-navn">{far_regnavn}</p>
                <p class="forelder-kallenavn">{far_under}</p>
            </div>
        </a>
        
    </div>

    <h2>Om kullet</h2>
    <p>{om_kullet}</p>
{valpe_seksjon}
    {BEVAR_START}{bevart}{BEVAR_SLUTT}

    <footer>
        {KENNEL_NAVN} · Følg oss på <a href="{FACEBOOK_URL}" target="_blank" rel="noopener">Facebook</a> for siste nytt
    </footer>

</div>

</body>
</html>
"""
    return html


# === HOVEDLOGIKK ===

def main():
    if not Path(KULL_EXCEL).exists():
        print(f"FEIL: Finner ikke {KULL_EXCEL}")
        return
    
    print(f"Leser {HUNDER_EXCEL}...")
    hunder = les_hunder(HUNDER_EXCEL)
    print(f"  {len(hunder)} hunder lastet.")
    
    print(f"\nLeser {KULL_EXCEL}...")
    alle_kull = les_kull(KULL_EXCEL)
    print(f"  {len(alle_kull)} kull funnet.\n")
    
    # Generer oversikt
    oversikt_html = generer_oversikt(alle_kull, hunder)
    Path(OVERSIKT_FIL).write_text(oversikt_html, encoding="utf-8")
    print(f"  Generert {OVERSIKT_FIL}")
    
    # Generer individuelle kullsider
    laget = 0
    oppdatert = 0
    for kull in alle_kull:
        filnavn = rens(kull.get("filnavn"))
        sti = Path(f"{filnavn}.html")
        eksisterte = sti.exists()
        
        html = generer_kullside(kull, hunder)
        sti.write_text(html, encoding="utf-8")
        
        if eksisterte:
            print(f"  Oppdaterte {sti.name}")
            oppdatert += 1
        else:
            print(f"  Laget {sti.name}")
            laget += 1
    
    print(f"\nFerdig!")
    print(f"  Nye filer: {laget}")
    print(f"  Oppdaterte filer: {oppdatert}")


if __name__ == "__main__":
    main()